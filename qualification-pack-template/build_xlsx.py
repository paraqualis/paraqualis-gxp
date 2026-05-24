#!/usr/bin/env python3
"""build_xlsx.py — assemble the qualification pack into one branded .xlsx
with a sheet per xQ stage plus Summary and Traceability sheets.

Markdown is the source of truth; this parses the tables out of docs/*.md.
    python3 build_xlsx.py   ->   docs/RegCheck-Qualification-Pack.xlsx
"""
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DOCS = Path(__file__).parent / "docs"
OUT = DOCS / "RegCheck-Qualification-Pack.xlsx"

DEEP_BLUE = "276BB2"
MINT = "E7F3EB"
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
CELL_FONT = Font(size=9)
WRAP = Alignment(vertical="top", wrap_text=True)
THIN = Border(*[Side(style="thin", color="D0D7DE")] * 4)


def table_after(md_path: Path, heading: str):
    """Return rows (list[list[str]]) of the first pipe-table after `heading`."""
    lines = md_path.read_text(encoding="utf-8").splitlines()
    out, capturing, seen_heading = [], False, False
    for ln in lines:
        if ln.strip().startswith("#") and heading.lower() in ln.lower():
            seen_heading = True; continue
        if seen_heading and ln.strip().startswith("|"):
            capturing = True
            if re.match(r"^\s*\|[\s:|-]+\|\s*$", ln):
                continue
            out.append([c.strip() for c in ln.strip().strip("|").split("|")])
        elif capturing and not ln.strip().startswith("|"):
            break
    return out


def write_sheet(wb, title, rows, first_sheet=False):
    ws = wb.active if first_sheet else wb.create_sheet()
    ws.title = title[:31]
    if not rows:
        ws["A1"] = "(no table found)"; return ws
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row, start=1):
            val = re.sub(r"\*\*", "", val)            # strip md bold markers
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = WRAP; cell.border = THIN
            if r == 1:
                cell.font = HDR_FONT
                cell.fill = PatternFill("solid", fgColor=DEEP_BLUE)
            else:
                cell.font = CELL_FONT
                if r % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=MINT)
    # column widths
    for c in range(1, len(rows[0]) + 1):
        width = max((len(str(row[c-1])) for row in rows if c-1 < len(row)), default=10)
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = min(max(width, 12), 60)
    ws.freeze_panes = "A2"
    return ws


if __name__ == "__main__":
    wb = Workbook()
    # Summary sheet: readiness verdict table
    write_sheet(wb, "Summary",
                table_after(DOCS / "Qualification-Summary.md", "Readiness verdict"),
                first_sheet=True)
    # One sheet per xQ stage: the Test-cases table
    write_sheet(wb, "IQ", table_after(DOCS / "IQ.md", "Test-cases"))
    write_sheet(wb, "OQ", table_after(DOCS / "OQ.md", "Test-cases"))
    write_sheet(wb, "PQ", table_after(DOCS / "PQ.md", "Test-cases"))
    # Traceability matrix
    write_sheet(wb, "Traceability",
                table_after(DOCS / "Qualification-Summary.md", "traceability matrix"))
    wb.save(str(OUT))
    print(f"wrote {OUT.name}  (sheets: {', '.join(s.title for s in wb.worksheets)})")
