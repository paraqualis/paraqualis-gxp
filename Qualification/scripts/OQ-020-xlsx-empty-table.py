#!/usr/bin/env python3
#
# Copyright (c) 2026 ParaQualis LLC
# Licensed under the MIT License — see LICENSE in the repository root.
#
"""OQ-020 — build_xlsx handles a missing heading / empty table gracefully.

Requirement (OQ.md OQ-020): table_after() returns [] when the heading is absent;
write_sheet() writes the '(no table found)' sentinel into A1 and does not crash.

Behaviour SOURCED from qualification-pack-template/build_xlsx.py:
    table_after() -> build_xlsx.py:25-39  (returns out; [] when heading never seen)
    write_sheet() -> build_xlsx.py:42-46  (if not rows: ws["A1"] = "(no table found)")

Read-only / safe: builds an in-memory Workbook only (never writes a file, never touches the
repo). Requires openpyxl (build_xlsx imports it at module load, build_xlsx.py:11). If openpyxl
is absent this FAILS LOUD with install guidance — it must never silently skip or print PASS
without running (GOQ-005 documents that the OQ test environment lacked openpyxl).
PASS/FAIL emitted; exits non-zero on FAIL.
"""
import importlib.util
import os
import sys
import tempfile
from pathlib import Path

BUILD_XLSX = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "..", "qualification-pack-template", "build_xlsx.py")
)

try:
    spec = importlib.util.spec_from_file_location("build_xlsx", BUILD_XLSX)
    build_xlsx = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(build_xlsx)
    from openpyxl import Workbook
except ModuleNotFoundError as e:
    print(f"FAIL: required dependency missing for build_xlsx ({e}). Run: pip install openpyxl "
          "(OQ-020 cannot be evidenced without it — GOQ-005)")
    sys.exit(1)

failures = []

# 1. table_after() returns [] when the heading is absent.
with tempfile.TemporaryDirectory() as d:
    md = Path(d) / "doc.md"
    md.write_text("# Some Other Heading\n\nNo table here at all.\n", encoding="utf-8")
    rows = build_xlsx.table_after(md, "Heading That Does Not Exist")
    if rows == []:
        print("PASS: table_after() returns [] for an absent heading")
    else:
        failures.append(f"table_after() returned {rows!r}, expected []")

# 2. write_sheet() with empty rows writes the sentinel and does not crash.
try:
    wb = Workbook()
    ws = build_xlsx.write_sheet(wb, "Empty", [], first_sheet=True)
    a1 = ws["A1"].value
    if a1 == "(no table found)":
        print("PASS: write_sheet([]) -> A1 == '(no table found)', no crash")
    else:
        failures.append(f"write_sheet([]) A1 == {a1!r}, expected '(no table found)'")
except Exception as e:
    failures.append(f"write_sheet([]) raised: {e!r}")

if failures:
    for f in failures:
        print(f"FAIL: {f}")
    sys.exit(1)
print("PASS: build_xlsx handles missing heading and empty table gracefully")
sys.exit(0)
